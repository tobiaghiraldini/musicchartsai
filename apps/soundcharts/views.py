from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
from .models import Track, Platform, Artist, TrackAudienceTimeSeries, ArtistAudienceTimeSeries, ChartRankingEntry, ChartSyncSchedule, ChartSyncExecution, Chart
from .audience_processor import AudienceDataProcessor
from .tasks import sync_chart_rankings_task
import json
import logging

logger = logging.getLogger(__name__)


@method_decorator(login_required, name='dispatch')
class TopArtistsView(View):

    def get(self, request):
        """
        Get top artists
        """
        return JsonResponse({
            'success': True,
            'artists': [
                {
                    'name': 'Artist 1',
                    'slug': 'artist-1',
                    'metric_name': 'Listeners'
                }
            ]
        })

@method_decorator(login_required, name='dispatch')
class TopSongsView(View):
    """
    View for getting top songs
    """
    
    def get(self, request):
        """
        Get top songs
        """
        return JsonResponse({
            'success': True,
            'songs': [
                {
                    'name': 'Song 1',
                    'slug': 'song-1',
                    'metric_name': 'Listeners'
                }
            ]
        })

@method_decorator(login_required, name="dispatch")
class PlatformListView(View):
    """
    View for listing platforms
    """
    
    def get(self, request):
        """
        Get all platforms
        """
        platforms = Platform.objects.all().values('name', 'slug', 'audience_metric_name')
        return JsonResponse({
            'success': True,
            'platforms': {platforms}
        })

@method_decorator(login_required, name='dispatch')
class AudienceChartView(View):
    """
    View for providing audience chart data
    """
    
    def get(self, request, track_uuid, platform_slug=None):
        """
        Get audience chart data for a track
        
        Args:
            track_uuid: SoundCharts track UUID
            platform_slug: Platform slug (optional, if not provided returns all platforms)
        """
        try:
            track = get_object_or_404(Track, uuid=track_uuid)
            
            if platform_slug:
                # Single platform chart data
                try:
                    platform = Platform.objects.get(slug=platform_slug)
                    chart_data = self._get_single_platform_data(track, platform, request)
                    return JsonResponse({
                        'success': True,
                        'track': {
                            'name': track.name,
                            'credit_name': track.credit_name,
                            'uuid': track.uuid
                        },
                        'platform': {
                            'name': platform.name,
                            'slug': platform.slug,
                            'metric_name': platform.audience_metric_name
                        },
                        'chart_data': chart_data
                    })
                except Platform.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Platform {platform_slug} not found'
                    }, status=404)
            else:
                # Multi-platform comparison data
                platforms = Platform.objects.filter(
                    audience_timeseries__track=track
                ).distinct()
                
                if not platforms.exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'No audience data found for this track'
                    }, status=404)
                
                chart_data = self._get_multi_platform_data(track, platforms, request)
                return JsonResponse({
                    'success': True,
                    'track': {
                        'name': track.name,
                        'credit_name': track.credit_name,
                        'uuid': track.uuid
                    },
                    'platforms': [{
                        'name': p.name,
                        'slug': p.slug,
                        'metric_name': p.audience_metric_name
                    } for p in platforms],
                    'chart_data': chart_data
                })
                
        except Exception as e:
            logger.error(f"Error in AudienceChartView: {e}")
            return JsonResponse({
                'success': False,
                'error': 'Internal server error'
            }, status=500)
    
    def _get_single_platform_data(self, track, platform, request):
        """Get chart data for a single platform"""
        # Get query parameters
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        limit = request.GET.get('limit')
        
        if limit:
            try:
                limit = int(limit)
            except ValueError:
                limit = None
        
        # Get the data using the model method
        try:
            data = track.get_audience_chart_data(
                platform, 
                start_date=start_date, 
                end_date=end_date, 
                limit=limit
            )
            # Convert to list to avoid queryset issues
            data = list(data)
            logger.info(f"Retrieved {len(data)} data points for {track.name} on {platform.name}")
        except Exception as e:
            logger.error(f"Error getting chart data for {track.name} on {platform.name}: {e}")
            data = []
        
        # Format for charting libraries
        chart_data = {
            'labels': [item['date'].strftime('%Y-%m-%d') for item in data],
            'datasets': [{
                'label': f'{track.name} on {platform.name}',
                'data': [item['audience_value'] for item in data],
                'borderColor': '#3b82f6',
                'backgroundColor': 'rgba(59, 130, 246, 0.1)',
                'tension': 0.1
            }]
        }
        
        return chart_data
    
    def _get_multi_platform_data(self, track, platforms, request):
        """Get chart data for multiple platforms comparison"""
        # Get query parameters
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        limit = request.GET.get('limit')
        
        if limit:
            try:
                limit = int(limit)
            except ValueError:
                limit = None
        
        # Get the data using the model method
        try:
            data = track.get_platform_audience_comparison(
                platforms, 
                start_date=start_date, 
                end_date=end_date, 
                limit=limit
            )
            # Convert to list to avoid queryset issues
            data = list(data)
            logger.info(f"Retrieved {len(data)} data points for {track.name} across {len(platforms)} platforms")
        except Exception as e:
            logger.error(f"Error getting platform comparison data for {track.name}: {e}")
            data = []
        
        # Group data by platform
        platform_data = {}
        dates = set()
        
        for item in data:
            platform_name = item['platform__name']
            date = item['date'].strftime('%Y-%m-%d')
            value = item['audience_value']
            
            if platform_name not in platform_data:
                platform_data[platform_name] = {}
            
            platform_data[platform_name][date] = value
            dates.add(date)
        
        # Sort dates
        sorted_dates = sorted(list(dates))
        
        # Create chart datasets
        datasets = []
        colors = ['#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6', '#ec4899']
        
        for i, (platform_name, date_values) in enumerate(platform_data.items()):
            color = colors[i % len(colors)]
            data_values = [date_values.get(date, None) for date in sorted_dates]
            
            datasets.append({
                'label': f'{track.name} on {platform_name}',
                'data': data_values,
                'borderColor': color,
                'backgroundColor': f'{color}20',
                'tension': 0.1
            })
        
        chart_data = {
            'labels': sorted_dates,
            'datasets': datasets
        }
        
        return chart_data


@method_decorator(login_required, name='dispatch')
class AudienceDataRefreshView(View):
    """
    View for manually refreshing audience data
    """
    
    def post(self, request, track_uuid, platform_slug):
        """
        Manually refresh audience data for a track on a specific platform
        """
        try:
            track = get_object_or_404(Track, uuid=track_uuid)
            platform = get_object_or_404(Platform, slug=platform_slug)
            
            # Process the data
            processor = AudienceDataProcessor()
            result = processor.process_and_store_audience_data(
                track_uuid, 
                platform_slug, 
                force_refresh=True
            )
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': f'Successfully refreshed audience data for {track.name} on {platform.name}',
                    'records_created': result['records_created'],
                    'records_updated': result['records_updated']
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result['error']
                }, status=400)
                
        except Exception as e:
            logger.error(f"Error in AudienceDataRefreshView: {e}")
            return JsonResponse({
                'success': False,
                'error': 'Internal server error'
            }, status=500)


# Legacy view for backward compatibility
@login_required
def audience_chart_data(request, track_uuid, platform_slug=None):
    """Legacy function-based view for audience chart data"""
    view = AudienceChartView()
    return view.get(request, track_uuid, platform_slug)


@method_decorator(login_required, name='dispatch')
class AudienceDashboardView(View):
    """
    View for the audience dashboard page
    """
    
    def get(self, request):
        """Render the audience dashboard page"""
        return render(request, 'dashboard/audience_charts.html')


@method_decorator(login_required, name='dispatch')
class TracksWithAudienceView(View):
    """
    API view to get tracks with audience data for the dashboard
    """
    
    def get(self, request):
        """
        Get all tracks that have audience data
        """
        try:
            # Get tracks that have audience time-series data
            tracks_with_audience = Track.objects.filter(
                audience_timeseries__isnull=False
            ).distinct().select_related('primary_artist').prefetch_related('audience_timeseries__platform')
            
            tracks_data = []
            
            for track in tracks_with_audience:
                # Get platforms for this track
                platforms = Platform.objects.filter(
                    audience_timeseries__track=track
                ).distinct()
                
                # Get latest audience data for each platform
                platforms_data = []
                for platform in platforms:
                    latest_audience = TrackAudienceTimeSeries.objects.filter(
                        track=track,
                        platform=platform
                    ).order_by('-date').first()
                    
                    platforms_data.append({
                        'name': platform.name,
                        'slug': platform.slug,
                        'metric_name': platform.audience_metric_name,
                        'latest_value': latest_audience.audience_value if latest_audience else 0,
                        'latest_date': latest_audience.date.isoformat() if latest_audience else None,
                        'data_points': TrackAudienceTimeSeries.objects.filter(
                            track=track,
                            platform=platform
                        ).count()
                    })
                
                # Include artist information
                artist_data = None
                if track.primary_artist:
                    artist_data = {
                        'uuid': track.primary_artist.uuid,
                        'name': track.primary_artist.name,
                        'slug': track.primary_artist.slug,
                    }
                
                tracks_data.append({
                    'uuid': track.uuid,
                    'name': track.name,
                    'credit_name': track.credit_name,
                    'image_url': track.image_url,
                    'artist': artist_data,
                    'audience_fetched_at': track.audience_fetched_at.isoformat() if track.audience_fetched_at else None,
                    'platforms': platforms_data
                })
            
            return JsonResponse({
                'success': True,
                'tracks': tracks_data,
                'total_tracks': len(tracks_data)
            })
            
        except Exception as e:
            logger.error(f"Error in TracksWithAudienceView: {e}")
            return JsonResponse({
                'success': False,
                'error': 'Internal server error'
            }, status=500)


@method_decorator(login_required, name='dispatch')
class SongAudienceDetailView(View):
    """
    View for displaying detailed audience analytics for a specific song
    """
    
    def get(self, request, track_uuid):
        """
        Display detailed audience analytics for a specific track
        """
        try:
            # Get the track
            track = get_object_or_404(Track, uuid=track_uuid)
            
            # Check if audience data needs to be fetched
            _check_and_fetch_audience_data(track)
            
            # Get platforms that have audience data for this track
            platforms = Platform.objects.filter(
                audience_timeseries__track=track
            ).distinct().order_by('name')
            
            # Get latest audience data for each platform
            platforms_data = []
            for platform in platforms:
                latest_audience = TrackAudienceTimeSeries.objects.filter(
                    track=track,
                    platform=platform
                ).order_by('-date').first()
                
                platforms_data.append({
                    'platform': platform,
                    'latest_audience': latest_audience,
                    'data_points': TrackAudienceTimeSeries.objects.filter(
                        track=track,
                        platform=platform
                    ).count()
                })
            
            # Get recent chart rankings for this track
            recent_rankings = ChartRankingEntry.objects.filter(
                track=track
            ).select_related('ranking__chart').order_by('-ranking__ranking_date')[:10]
            
            # Get related tracks (same artist or similar genres)
            related_tracks = []
            if track.primary_artist:
                related_tracks = Track.objects.filter(
                    primary_artist=track.primary_artist
                ).exclude(uuid=track.uuid)[:5]
            elif track.artists.exists():
                related_tracks = Track.objects.filter(
                    artists__in=track.artists.all()
                ).exclude(uuid=track.uuid).distinct()[:5]
            
            context = {
                'track': track,
                'platforms_data': platforms_data,
                'recent_rankings': recent_rankings,
                'related_tracks': related_tracks,
                'segment': 'charts',
            }
            
            return render(request, 'soundcharts/song_audience_detail.html', context)
            
        except Exception as e:
            logger.error(f"Error in SongAudienceDetailView: {e}")
            return render(request, 'soundcharts/song_audience_detail.html', {
                'error': 'Failed to load song details',
                'segment': 'charts',
            })


# Chart Sync API Views
@method_decorator(staff_member_required, name='dispatch')
class ChartSyncScheduleAPIView(View):
    """
    API view for managing chart sync schedules
    """
    
    def get(self, request):
        """Get all chart sync schedules"""
        try:
            schedules = ChartSyncSchedule.objects.select_related('chart', 'chart__platform').all()
            
            schedules_data = []
            for schedule in schedules:
                schedules_data.append({
                    'id': schedule.id,
                    'chart': {
                        'id': schedule.chart.id,
                        'name': schedule.chart.name,
                        'platform': schedule.chart.platform.name if schedule.chart.platform else None,
                        'frequency': schedule.chart.frequency,
                    },
                    'is_active': schedule.is_active,
                    'sync_frequency': schedule.sync_frequency,
                    'custom_interval_hours': schedule.custom_interval_hours,
                    'last_sync_at': schedule.last_sync_at.isoformat() if schedule.last_sync_at else None,
                    'next_sync_at': schedule.next_sync_at.isoformat() if schedule.next_sync_at else None,
                    'total_executions': schedule.total_executions,
                    'successful_executions': schedule.successful_executions,
                    'failed_executions': schedule.failed_executions,
                    'success_rate': schedule.success_rate,
                    'is_overdue': schedule.is_overdue,
                    'created_at': schedule.created_at.isoformat(),
                })
            
            return JsonResponse({
                'success': True,
                'schedules': schedules_data
            })
            
        except Exception as e:
            logger.error(f"Error getting chart sync schedules: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    def post(self, request):
        """Create a new chart sync schedule"""
        try:
            data = json.loads(request.body)
            chart_id = data.get('chart_id')
            
            if not chart_id:
                return JsonResponse({
                    'success': False,
                    'error': 'chart_id is required'
                }, status=400)
            
            chart = get_object_or_404(Chart, id=chart_id)
            
            # Check if schedule already exists
            if ChartSyncSchedule.objects.filter(chart=chart).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Chart sync schedule already exists for this chart'
                }, status=400)
            
            # Create new schedule
            schedule = ChartSyncSchedule.objects.create(
                chart=chart,
                created_by=request.user,
                is_active=data.get('is_active', True),
                sync_frequency=data.get('sync_frequency'),
                custom_interval_hours=data.get('custom_interval_hours'),
                sync_immediately=data.get('sync_immediately', False),
                sync_historical_data=data.get('sync_historical_data', True),
                fetch_track_metadata=data.get('fetch_track_metadata', True),
            )
            
            return JsonResponse({
                'success': True,
                'schedule': {
                    'id': schedule.id,
                    'chart': {
                        'id': schedule.chart.id,
                        'name': schedule.chart.name,
                    },
                    'is_active': schedule.is_active,
                    'sync_frequency': schedule.sync_frequency,
                    'custom_interval_hours': schedule.custom_interval_hours,
                    'sync_immediately': schedule.sync_immediately,
                    'sync_historical_data': schedule.sync_historical_data,
                    'fetch_track_metadata': schedule.fetch_track_metadata,
                    'next_sync_at': schedule.next_sync_at.isoformat() if schedule.next_sync_at else None,
                }
            })
            
        except Exception as e:
            logger.error(f"Error creating chart sync schedule: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(staff_member_required, name='dispatch')
class ChartSyncScheduleDetailAPIView(View):
    """
    API view for managing individual chart sync schedules
    """
    
    def get(self, request, schedule_id):
        """Get a specific chart sync schedule"""
        try:
            schedule = get_object_or_404(ChartSyncSchedule, id=schedule_id)
            
            # Get recent executions
            executions = ChartSyncExecution.objects.filter(
                schedule=schedule
            ).order_by('-started_at')[:10]
            
            executions_data = []
            for execution in executions:
                executions_data.append({
                    'id': execution.id,
                    'status': execution.status,
                    'started_at': execution.started_at.isoformat(),
                    'completed_at': execution.completed_at.isoformat() if execution.completed_at else None,
                    'duration': execution.duration,
                    'rankings_created': execution.rankings_created,
                    'rankings_updated': execution.rankings_updated,
                    'tracks_created': execution.tracks_created,
                    'tracks_updated': execution.tracks_updated,
                    'error_message': execution.error_message,
                })
            
            return JsonResponse({
                'success': True,
                'schedule': {
                    'id': schedule.id,
                    'chart': {
                        'id': schedule.chart.id,
                        'name': schedule.chart.name,
                        'platform': schedule.chart.platform.name if schedule.chart.platform else None,
                        'frequency': schedule.chart.frequency,
                    },
                    'is_active': schedule.is_active,
                    'sync_frequency': schedule.sync_frequency,
                    'custom_interval_hours': schedule.custom_interval_hours,
                    'last_sync_at': schedule.last_sync_at.isoformat() if schedule.last_sync_at else None,
                    'next_sync_at': schedule.next_sync_at.isoformat() if schedule.next_sync_at else None,
                    'total_executions': schedule.total_executions,
                    'successful_executions': schedule.successful_executions,
                    'failed_executions': schedule.failed_executions,
                    'success_rate': schedule.success_rate,
                    'is_overdue': schedule.is_overdue,
                    'created_at': schedule.created_at.isoformat(),
                    'executions': executions_data,
                }
            })
            
        except Exception as e:
            logger.error(f"Error getting chart sync schedule {schedule_id}: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    def put(self, request, schedule_id):
        """Update a chart sync schedule"""
        try:
            schedule = get_object_or_404(ChartSyncSchedule, id=schedule_id)
            data = json.loads(request.body)
            
            # Update fields
            if 'is_active' in data:
                schedule.is_active = data['is_active']
            if 'sync_frequency' in data:
                schedule.sync_frequency = data['sync_frequency']
            if 'custom_interval_hours' in data:
                schedule.custom_interval_hours = data['custom_interval_hours']
            
            schedule.save()
            
            return JsonResponse({
                'success': True,
                'schedule': {
                    'id': schedule.id,
                    'is_active': schedule.is_active,
                    'sync_frequency': schedule.sync_frequency,
                    'custom_interval_hours': schedule.custom_interval_hours,
                    'next_sync_at': schedule.next_sync_at.isoformat() if schedule.next_sync_at else None,
                }
            })
            
        except Exception as e:
            logger.error(f"Error updating chart sync schedule {schedule_id}: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    def delete(self, request, schedule_id):
        """Delete a chart sync schedule"""
        try:
            schedule = get_object_or_404(ChartSyncSchedule, id=schedule_id)
            chart_name = schedule.chart.name
            schedule.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Chart sync schedule for "{chart_name}" deleted successfully'
            })
            
        except Exception as e:
            logger.error(f"Error deleting chart sync schedule {schedule_id}: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(staff_member_required, name='dispatch')
class ChartSyncTriggerAPIView(View):
    """
    API view for triggering manual chart syncs
    """
    
    def post(self, request):
        """Trigger manual sync for a chart"""
        try:
            data = json.loads(request.body)
            chart_id = data.get('chart_id')
            
            if not chart_id:
                return JsonResponse({
                    'success': False,
                    'error': 'chart_id is required'
                }, status=400)
            
            chart = get_object_or_404(Chart, id=chart_id)
            
            try:
                schedule = ChartSyncSchedule.objects.get(chart=chart)
            except ChartSyncSchedule.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Chart sync schedule not found for this chart'
                }, status=404)
            
            if not schedule.is_active:
                return JsonResponse({
                    'success': False,
                    'error': 'Chart sync schedule is not active'
                }, status=400)
            
            # Create execution record
            execution = ChartSyncExecution.objects.create(
                schedule=schedule,
                status='pending'
            )
            
            # Queue the sync task
            task = sync_chart_rankings_task.delay(schedule.id, execution.id)
            execution.celery_task_id = task.id
            execution.status = 'running'
            execution.save()
            
            return JsonResponse({
                'success': True,
                'execution': {
                    'id': execution.id,
                    'status': execution.status,
                    'celery_task_id': execution.celery_task_id,
                    'started_at': execution.started_at.isoformat(),
                },
                'message': f'Manual sync triggered for chart "{chart.name}"'
            })
            
        except Exception as e:
            logger.error(f"Error triggering manual sync: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(staff_member_required, name='dispatch')
class ChartSyncStatusAPIView(View):
    """
    API view for getting chart sync status
    """
    
    def get(self, request, chart_id):
        """Get sync status for a specific chart"""
        try:
            chart = get_object_or_404(Chart, id=chart_id)
            
            try:
                schedule = ChartSyncSchedule.objects.get(chart=chart)
                
                # Get latest execution
                latest_execution = ChartSyncExecution.objects.filter(
                    schedule=schedule
                ).order_by('-started_at').first()
                
                return JsonResponse({
                    'success': True,
                    'status': {
                        'has_schedule': True,
                        'is_active': schedule.is_active,
                        'sync_frequency': schedule.sync_frequency,
                        'last_sync_at': schedule.last_sync_at.isoformat() if schedule.last_sync_at else None,
                        'next_sync_at': schedule.next_sync_at.isoformat() if schedule.next_sync_at else None,
                        'total_executions': schedule.total_executions,
                        'success_rate': schedule.success_rate,
                        'is_overdue': schedule.is_overdue,
                        'latest_execution': {
                            'id': latest_execution.id,
                            'status': latest_execution.status,
                            'started_at': latest_execution.started_at.isoformat(),
                            'completed_at': latest_execution.completed_at.isoformat() if latest_execution.completed_at else None,
                            'error_message': latest_execution.error_message,
                        } if latest_execution else None,
                    }
                })
                
            except ChartSyncSchedule.DoesNotExist:
                return JsonResponse({
                    'success': True,
                    'status': {
                        'has_schedule': False,
                        'is_active': False,
                    }
                })
            
        except Exception as e:
            logger.error(f"Error getting chart sync status for chart {chart_id}: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


def _check_and_fetch_audience_data(track):
    """
    Check if audience data needs to be fetched and trigger fetch if needed
    """
    from datetime import timedelta
    from django.utils import timezone
    from .tasks import fetch_track_audience_data
    
    # Check if audience data is stale or missing
    should_fetch = False
    
    if not track.audience_fetched_at:
        # Never fetched audience data
        should_fetch = True
    else:
        # Check if audience data is older than 7 days
        cutoff_date = timezone.now() - timedelta(days=7)
        if track.audience_fetched_at < cutoff_date:
            should_fetch = True
    
    # Check if track has any audience data at all
    if not TrackAudienceTimeSeries.objects.filter(track=track).exists():
        should_fetch = True
    
    if should_fetch:
        # Queue audience data fetch task
        fetch_track_audience_data.delay(track.uuid)
        logger.info(f"Queued audience data fetch for track {track.uuid}")


# ============================================================================
# ARTIST VIEWS
# ============================================================================

@method_decorator(login_required, name='dispatch')
class ArtistSearchView(View):
    """View for searching artists via Soundcharts API"""
    
    def get(self, request):
        """Render the artist search page"""
        return render(request, 'soundcharts/artist_search.html', {
            'segment': 'artists',
        })
    
    def post(self, request):
        """Search for artists via API"""
        try:
            data = json.loads(request.body)
            query = data.get('query', '').strip()
            
            if not query:
                return JsonResponse({
                    'success': False,
                    'error': 'Search query is required'
                }, status=400)
            
            from .service import SoundchartsService
            service = SoundchartsService()
            
            # Search for artists
            results = service.search_artists(query, limit=20)
            
            if results:
                return JsonResponse({
                    'success': True,
                    'artists': results
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'No artists found or API error occurred'
                })
                
        except Exception as e:
            logger.error(f"Error searching artists: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(login_required, name='dispatch')
class ArtistListView(View):
    """View for listing stored artists"""
    
    def get(self, request):
        """Get all stored artists with their audience data"""
        try:
            artists = Artist.objects.all().prefetch_related(
                'genres', 
                'audience_timeseries__platform'
            ).order_by('name')
            
            artists_data = []
            for artist in artists:
                # Get platforms that have audience data
                platforms_with_data = Platform.objects.filter(
                    artist_audience_timeseries__artist=artist
                ).distinct()
                
                artists_data.append({
                    'id': artist.id,
                    'uuid': artist.uuid,
                    'name': artist.name,
                    'slug': artist.slug,
                    'image_url': artist.imageUrl,
                    'country_code': artist.countryCode,
                    'career_stage': artist.careerStage,
                    'has_audience_data': platforms_with_data.exists(),
                    'platforms_count': platforms_with_data.count(),
                    'metadata_fetched': artist.metadata_fetched_at is not None,
                    'audience_fetched': artist.audience_fetched_at is not None,
                })
            
            return render(request, 'soundcharts/artist_list.html', {
                'artists': artists_data,
                'segment': 'artists',
            })
            
        except Exception as e:
            logger.error(f"Error in ArtistListView: {e}")
            return render(request, 'soundcharts/artist_list.html', {
                'error': 'Failed to load artists',
                'segment': 'artists',
            })


@method_decorator(login_required, name='dispatch')
class ArtistDetailView(View):
    """View for displaying detailed analytics for a specific artist"""
    
    def get(self, request, artist_uuid):
        """Display detailed audience analytics for a specific artist"""
        try:
            # Get the artist
            artist = get_object_or_404(Artist, uuid=artist_uuid)
            
            # Get platforms that have audience data for this artist
            platforms = Platform.objects.filter(
                artist_audience_timeseries__artist=artist
            ).distinct().order_by('name')
            
            # Get latest audience data for each platform
            platforms_data = []
            for platform in platforms:
                latest_audience = ArtistAudienceTimeSeries.objects.filter(
                    artist=artist,
                    platform=platform
                ).order_by('-date').first()
                
                platforms_data.append({
                    'platform': platform,
                    'latest_audience': latest_audience,
                    'data_points': ArtistAudienceTimeSeries.objects.filter(
                        artist=artist,
                        platform=platform
                    ).count()
                })
            
            # Get tracks by this artist
            related_tracks = Track.objects.filter(
                Q(primary_artist=artist) | Q(artists=artist)
            ).distinct()[:10]
            
            context = {
                'artist': artist,
                'platforms_data': platforms_data,
                'related_tracks': related_tracks,
                'segment': 'artists',
            }
            
            return render(request, 'soundcharts/artist_detail.html', context)
            
        except Exception as e:
            logger.error(f"Error in ArtistDetailView: {e}")
            return render(request, 'soundcharts/artist_detail.html', {
                'error': 'Failed to load artist details',
                'segment': 'artists',
            })


@method_decorator(login_required, name='dispatch')
class ArtistAudienceChartView(View):
    """View for providing artist audience chart data"""
    
    def get(self, request, artist_uuid, platform_slug=None):
        """Get audience chart data for an artist"""
        try:
            artist = get_object_or_404(Artist, uuid=artist_uuid)
            
            if platform_slug:
                # Single platform chart data
                try:
                    platform = Platform.objects.get(slug=platform_slug)
                    chart_data = self._get_single_platform_data(artist, platform, request)
                    return JsonResponse({
                        'success': True,
                        'artist': {
                            'name': artist.name,
                            'uuid': artist.uuid
                        },
                        'platform': {
                            'name': platform.name,
                            'slug': platform.slug,
                            'metric_name': platform.audience_metric_name
                        },
                        'chart_data': chart_data
                    })
                except Platform.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Platform {platform_slug} not found'
                    }, status=404)
            else:
                # All platforms chart data
                platforms = Platform.objects.filter(
                    artist_audience_timeseries__artist=artist
                ).distinct()
                
                platforms_data = []
                for platform in platforms:
                    chart_data = self._get_single_platform_data(artist, platform, request)
                    platforms_data.append({
                        'platform': {
                            'name': platform.name,
                            'slug': platform.slug,
                            'metric_name': platform.audience_metric_name
                        },
                        'chart_data': chart_data
                    })
                
                return JsonResponse({
                    'success': True,
                    'artist': {
                        'name': artist.name,
                        'uuid': artist.uuid
                    },
                    'platforms': platforms_data
                })
                
        except Exception as e:
            logger.error(f"Error in ArtistAudienceChartView: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    def _get_single_platform_data(self, artist, platform, request):
        """Get chart data for a single platform"""
        try:
            limit = int(request.GET.get('limit', 30))
            
            # Get time-series data
            chart_data = ArtistAudienceTimeSeries.get_chart_data(
                artist=artist,
                platform=platform,
                limit=limit
            )
            
            if not chart_data:
                return {
                    'labels': [],
                    'datasets': [{
                        'label': platform.audience_metric_name,
                        'data': []
                    }]
                }
            
            # Convert to chart.js format
            labels = [str(item['date']) for item in chart_data]
            values = [item['audience_value'] for item in chart_data]
            
            return {
                'labels': labels,
                'datasets': [{
                    'label': platform.audience_metric_name,
                    'data': values,
                    'fill': True,
                    'tension': 0.4
                }]
            }
            
        except Exception as e:
            logger.error(f"Error getting chart data for {artist.name} on {platform.slug}: {e}")
            return {
                'labels': [],
                'datasets': [{
                    'label': platform.audience_metric_name,
                    'data': []
                }]
            }


@method_decorator(login_required, name='dispatch')
class ArtistSaveView(View):
    """API view to save an artist from search results to database"""
    
    def post(self, request):
        """Save an artist from Soundcharts to database"""
        try:
            data = json.loads(request.body)
            artist_data = data.get('artist')
            
            if not artist_data:
                return JsonResponse({
                    'success': False,
                    'error': 'No artist data provided'
                }, status=400)
            
            # Create or update artist
            artist = Artist.create_from_soundcharts(artist_data)
            
            if artist:
                return JsonResponse({
                    'success': True,
                    'message': f'Artist "{artist.name}" saved successfully',
                    'artist': {
                        'id': artist.id,
                        'uuid': artist.uuid,
                        'name': artist.name,
                        'slug': artist.slug
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to save artist'
                }, status=500)
                
        except Exception as e:
            logger.error(f"Error saving artist: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(login_required, name='dispatch')
class ArtistMetadataFetchView(View):
    """API view to fetch and update artist metadata"""
    
    def post(self, request):
        """Fetch metadata for an artist"""
        try:
            data = json.loads(request.body)
            artist_uuid = data.get('uuid')
            
            if not artist_uuid:
                return JsonResponse({
                    'success': False,
                    'error': 'Artist UUID is required'
                }, status=400)
            
            # Get or create artist
            artist = Artist.objects.filter(uuid=artist_uuid).first()
            if not artist:
                return JsonResponse({
                    'success': False,
                    'error': 'Artist not found in database'
                }, status=404)
            
            # Fetch metadata from API
            from .service import SoundchartsService
            service = SoundchartsService()
            metadata = service.get_artist_metadata(artist_uuid)
            
            if metadata and 'object' in metadata:
                artist_data = metadata['object']
                
                # Update artist fields
                if 'name' in artist_data:
                    artist.name = artist_data['name']
                if 'biography' in artist_data:
                    artist.biography = artist_data['biography']
                if 'countryCode' in artist_data:
                    artist.countryCode = artist_data['countryCode']
                if 'careerStage' in artist_data:
                    artist.careerStage = artist_data['careerStage']
                # Add more fields as needed
                
                artist.metadata_fetched_at = timezone.now()
                artist.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Metadata fetched for "{artist.name}"'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to fetch metadata from API'
                }, status=500)
                
        except Exception as e:
            logger.error(f"Error fetching artist metadata: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


@method_decorator(login_required, name='dispatch')
class ArtistAudienceFetchView(View):
    """API view to fetch and store artist audience data"""
    
    def post(self, request):
        """Fetch audience data for an artist"""
        try:
            data = json.loads(request.body)
            artist_uuid = data.get('uuid')
            platform_slug = data.get('platform', 'spotify')
            
            if not artist_uuid:
                return JsonResponse({
                    'success': False,
                    'error': 'Artist UUID is required'
                }, status=400)
            
            # Get artist
            artist = Artist.objects.filter(uuid=artist_uuid).first()
            if not artist:
                return JsonResponse({
                    'success': False,
                    'error': 'Artist not found in database'
                }, status=404)
            
            # Get platform
            platform = Platform.objects.filter(slug=platform_slug).first()
            if not platform:
                return JsonResponse({
                    'success': False,
                    'error': f'Platform {platform_slug} not found'
                }, status=404)
            
            # Fetch audience data from API
            from .service import SoundchartsService
            service = SoundchartsService()
            audience_data = service.get_artist_audience_for_platform(artist_uuid, platform=platform_slug)
            
            if audience_data and 'items' in audience_data:
                # Process and store time-series data
                items = audience_data.get('items', [])
                records_created = 0
                records_updated = 0
                
                for item in items:
                    item_date_str = item.get('date')
                    if not item_date_str:
                        continue
                    
                    try:
                        # Parse date
                        if 'T' in item_date_str:
                            item_date = datetime.fromisoformat(item_date_str.replace('Z', '+00:00')).date()
                        else:
                            item_date = datetime.strptime(item_date_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        continue
                    
                    # Get audience value
                    audience_value = (item.get('followerCount') or 
                                    item.get('likeCount') or 
                                    item.get('viewCount'))
                    if audience_value is None:
                        continue
                    
                    # Store time-series record
                    ts_record, created = ArtistAudienceTimeSeries.objects.update_or_create(
                        artist=artist,
                        platform=platform,
                        date=item_date,
                        defaults={
                            'audience_value': audience_value,
                            'api_data': item,
                            'fetched_at': timezone.now()
                        }
                    )
                    
                    if created:
                        records_created += 1
                    else:
                        records_updated += 1
                
                # Update fetch timestamp
                artist.audience_fetched_at = timezone.now()
                artist.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Fetched audience data for "{artist.name}" on {platform.name}: {records_created} created, {records_updated} updated',
                    'data': {
                        'records_created': records_created,
                        'records_updated': records_updated
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to fetch audience data from API'
                }, status=500)
                
        except Exception as e:
            logger.error(f"Error fetching artist audience: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


# ============================================================================
# MUSIC ANALYTICS VIEWS (Phase 1: Artist-Level Aggregation)
# ============================================================================

@login_required
def analytics_search_form(request):
    """
    Display the music analytics search form.
    
    This form allows users to select artists, platforms, and date range
    to aggregate audience metrics.
    
    Query parameters:
    - artist_uuid: Pre-select an artist by UUID (for deep linking from artist detail page)
    
    TODO: Add role-based permissions in the future:
    - @permission_required('soundcharts.view_analytics')
    - Different user roles: admin, analyst, viewer
    - Limit date range access based on subscription tier
    - Rate limiting for expensive aggregation queries
    """
    from .analytics_service import MusicAnalyticsService
    import json
    
    service = MusicAnalyticsService()
    
    # Get supported platforms for analytics
    # Only include platforms that work with SoundCharts streaming/social endpoints
    supported_slugs = ['spotify', 'youtube', 'tiktok', 'shazam', 'airplay']
    platforms = Platform.objects.filter(
        slug__in=supported_slugs
    ).order_by('name')
    
    # Log available platforms for debugging
    logger.info(f"Analytics form: {platforms.count()} supported platforms available")
    
    # Get available countries (informational only)
    countries = service.get_available_countries()
    
    # Check for artist_uuid query parameter (for deep linking from artist detail page)
    artist_uuid = request.GET.get('artist_uuid', '').strip()
    prefill_artist = None
    
    if artist_uuid:
        try:
            from .models import Artist
            artist = Artist.objects.filter(uuid=artist_uuid).first()
            if artist:
                prefill_artist = {
                    'id': artist.id,
                    'uuid': artist.uuid,
                    'name': artist.name,
                    'imageUrl': artist.imageUrl or '',
                    'countryCode': artist.countryCode or '',
                    'has_uuid': bool(artist.uuid)
                }
        except Exception as e:
            logger.warning(f"Error looking up artist by UUID {artist_uuid}: {e}")
    
    context = {
        'platforms': platforms,
        'countries': countries,
        'prefill_artist_json': json.dumps(prefill_artist) if prefill_artist else 'null',
    }
    
    return render(request, 'soundcharts/analytics_search.html', context)


@login_required
def analytics_search_results(request):
    """
    Process analytics search and display results.
    
    Handles both GET (initial search) and POST (refresh/export) requests.
    """
    from .analytics_service import MusicAnalyticsService
    from datetime import datetime
    
    service = MusicAnalyticsService()
    
    if request.method != 'POST':
        # Redirect to search form if accessed via GET
        from django.shortcuts import redirect
        return redirect('soundcharts:analytics_search')
    
    try:
        # Parse request data
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        
        artist_ids = data.getlist('artist_ids[]') if hasattr(data, 'getlist') else data.get('artist_ids', [])
        platform_ids = data.getlist('platform_ids[]') if hasattr(data, 'getlist') else data.get('platform_ids', [])
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        country = data.get('country', '')  # Informational only
        
        # Validate inputs
        if not artist_ids:
            return JsonResponse({
                'success': False,
                'error': 'Please select at least one artist'
            }, status=400)
        
        if not platform_ids:
            return JsonResponse({
                'success': False,
                'error': 'Please select at least one platform'
            }, status=400)
        
        if not start_date_str or not end_date_str:
            return JsonResponse({
                'success': False,
                'error': 'Please select a date range'
            }, status=400)
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate date range
        if end_date < start_date:
            return JsonResponse({
                'success': False,
                'error': 'End date must be after start date'
            }, status=400)
        
        date_range_days = (end_date - start_date).days + 1
        if date_range_days > 365:
            return JsonResponse({
                'success': False,
                'error': 'Date range cannot exceed 365 days'
            }, status=400)
        
        # Fetch fresh data from SoundCharts API and aggregate
        # This is the main Phase 1 functionality - fetch data on demand
        result = service.fetch_and_aggregate_artist_metrics(
            artist_ids, platform_ids, start_date, end_date, country
        )
        
        if not result['success']:
            # Include detailed error info for debugging
            error_response = {
                'success': False,
                'error': result['error']
            }
            if 'error_details' in result:
                error_response['error_details'] = result['error_details']
            
            logger.error(f"Analytics search failed: {result.get('error')}")
            if 'error_details' in result:
                logger.error(f"Error details: {result['error_details']}")
            
            return JsonResponse(error_response, status=400)
        
        # Add country info (informational)
        result['data']['metadata']['country'] = country
        
        # Store search params for Excel export
        result['data']['search_params'] = {
            'artist_ids': artist_ids,
            'platform_ids': platform_ids,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'country': country
        }
        
        # For AJAX requests, return JSON with results
        # The frontend will handle displaying the results
        return JsonResponse(result)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'Invalid date format: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.error(f"Error in analytics search: {e}")
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)


@login_required
def analytics_artist_autocomplete(request):
    """
    Autocomplete endpoint for artist selection in analytics form.
    
    Only returns artists with valid SoundCharts UUIDs, as these are required
    for fetching audience data from the API.
    
    NOTE: Artists from ACRCloud analysis may not have SoundCharts UUIDs.
    These artists cannot be used for audience aggregation until their
    SoundCharts UUID is populated via manual search or API matching.
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({
            'success': True,
            'artists': []
        })
    
    # Only return artists with valid UUIDs
    artists = Artist.objects.filter(
        name__icontains=query,
        uuid__isnull=False
    ).exclude(uuid='').order_by('name')[:20]
    
    results = [{
        'id': artist.id,
        'uuid': artist.uuid,
        'name': artist.name,
        'imageUrl': artist.imageUrl,
        'countryCode': artist.countryCode or '',
        'has_uuid': bool(artist.uuid)
    } for artist in artists]
    
    return JsonResponse({
        'success': True,
        'artists': results
    })


@login_required
def analytics_track_breakdown(request):
    """
    Get track-level breakdown for a specific artist × platform combination.
    
    Phase 2: Returns track streaming data for expandable table rows.
    
    Query params:
        - artist_id: Artist ID
        - platform_id: Platform ID
        - start_date: Start date (YYYY-MM-DD)
        - end_date: End date (YYYY-MM-DD)
        - country: Optional country code
    """
    from .analytics_service import MusicAnalyticsService
    from datetime import datetime
    
    service = MusicAnalyticsService()
    
    try:
        artist_id = request.GET.get('artist_id')
        platform_id = request.GET.get('platform_id')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        country = request.GET.get('country', '')
        
        if not all([artist_id, platform_id, start_date_str, end_date_str]):
            return JsonResponse({
                'success': False,
                'error': 'Missing required parameters'
            }, status=400)
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Get track breakdown
        result = service.get_track_breakdown_for_artist(
            int(artist_id), int(platform_id), start_date, end_date, country
        )
        
        return JsonResponse(result)
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'Invalid parameter format: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.error(f"Error fetching track breakdown: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)


@login_required
def analytics_export_excel(request):
    """
    Export analytics results to Excel file.
    
    Simple flat table format with metadata header.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    from .analytics_service import MusicAnalyticsService
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'POST request required'
        }, status=400)
    
    try:
        # Get the same data as the results page
        service = MusicAnalyticsService()
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        
        # Parse parameters (handle both list and single values)
        artist_ids = data.get('artist_ids', [])
        if not isinstance(artist_ids, list):
            artist_ids = [artist_ids]
        
        platform_ids = data.get('platform_ids', [])
        if not isinstance(platform_ids, list):
            platform_ids = [platform_ids]
        
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        country = data.get('country', '')
        
        logger.info(f"Excel export request: artists={artist_ids}, platforms={platform_ids}, dates={start_date_str} to {end_date_str}, country={country}")
        
        # Validate required fields
        if not artist_ids or not platform_ids or not start_date_str or not end_date_str:
            return JsonResponse({
                'success': False,
                'error': 'Missing required parameters for export'
            }, status=400)
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Fetch fresh data (same as results page)
        result = service.fetch_and_aggregate_artist_metrics(
            artist_ids, platform_ids, start_date, end_date, country
        )
        
        if not result['success']:
            return JsonResponse({
                'success': False,
                'error': result['error']
            }, status=400)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Metadata section
        ws['A1'] = "Music Analytics Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Date Range: {start_date} to {end_date}"
        ws['A4'] = f"Artists: {', '.join([a['name'] for a in result['data']['metadata']['artists']])}"
        ws['A5'] = f"Platforms: {', '.join([p['name'] for p in result['data']['metadata']['platforms']])}"
        ws['A6'] = f"Country: {country if country else 'Global (All Countries)'}"
        
        # Per-Platform Summary section (from the colored cards)
        row = 8
        if result['data'].get('platform_summaries'):
            ws[f'A{row}'] = "PER-PLATFORM SUMMARY"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Headers for platform summary
            platform_headers = ['Platform', 'Metric Type', 'Start Value', 'End Value', 'Difference', 'Period Average', 'Peak Value', 'Data Points']
            for col, header in enumerate(platform_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Platform summary data
            for platform_summary in result['data']['platform_summaries']:
                row += 1
                ws.cell(row=row, column=1, value=platform_summary['platform__name'])
                ws.cell(row=row, column=2, value=platform_summary.get('platform__audience_metric_name', 'Monthly Listeners'))
                ws.cell(row=row, column=3, value=platform_summary.get('first_value', 0))
                ws.cell(row=row, column=4, value=platform_summary.get('latest_value', 0))
                ws.cell(row=row, column=5, value=platform_summary.get('difference', 0))
                ws.cell(row=row, column=6, value=platform_summary.get('period_average', 0))
                ws.cell(row=row, column=7, value=platform_summary.get('peak_value', 0))
                ws.cell(row=row, column=8, value=platform_summary.get('data_points', 0))
            
            row += 2
        
        # Detailed breakdown section (Artist x Platform)
        ws[f'A{row}'] = "ARTIST x PLATFORM BREAKDOWN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        headers = ['Artist', 'Platform', 'Metric Type', 'Start Value', 'End Value', 'Difference', 'Period Average', 'Peak Value', 'Track Streams', 'Data Points']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for detail in result['data']['detailed_breakdown']:
            row += 1
            # Determine metric type
            metric_type = detail.get('platform__audience_metric_name', 'Monthly Listeners')
            
            ws.cell(row=row, column=1, value=detail['artist__name'])
            ws.cell(row=row, column=2, value=detail['platform__name'])
            ws.cell(row=row, column=3, value=metric_type)
            ws.cell(row=row, column=4, value=detail.get('first_value', 0))
            ws.cell(row=row, column=5, value=detail.get('latest_value', 0))
            ws.cell(row=row, column=6, value=detail.get('difference', 0))
            ws.cell(row=row, column=7, value=detail.get('month_average', 0))
            ws.cell(row=row, column=8, value=detail.get('peak_value', 0))
            ws.cell(row=row, column=9, value=detail.get('total_track_streams', 0))
            ws.cell(row=row, column=10, value=detail.get('data_points', 0))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # NEW: Add track breakdown sheets for each artist × platform
        logger.info("Adding track breakdown sheets...")
        artists = result['data']['metadata']['artists']
        platforms = result['data']['metadata']['platforms']
        
        for artist in artists:
            for platform in platforms:
                # Fetch track breakdown data
                track_data = service.get_track_breakdown_for_artist(
                    artist['id'], 
                    platform['id'], 
                    start_date, 
                    end_date, 
                    country
                )
                
                if not track_data.get('success') or not track_data.get('tracks'):
                    logger.info(f"No track data for {artist['name']} on {platform['name']}")
                    continue
                
                # Create new sheet for this artist × platform
                sheet_name = f"{artist['name'][:20]} - {platform['name'][:10]}"  # Excel sheet name limit: 31 chars
                track_ws = wb.create_sheet(title=sheet_name)
                
                # Header section
                track_ws['A1'] = f"Track Breakdown: {artist['name']} on {platform['name']}"
                track_ws['A1'].font = Font(bold=True, size=12)
                track_ws['A2'] = f"Period: {start_date} to {end_date}"
                track_ws['A3'] = f"Country: {country if country else 'Global'}"
                
                # Summary metrics
                track_ws['A5'] = "Total Tracks:"
                track_ws['B5'] = track_data['summary']['total_tracks']
                track_ws['B5'].font = Font(bold=True)
                
                track_ws['A6'] = "Total Streams:"
                track_ws['B6'] = track_data['summary']['total_streams']
                track_ws['B6'].font = Font(bold=True)
                track_ws['C6'] = service.format_number(track_data['summary']['total_streams'])
                
                track_ws['A7'] = "Avg Streams/Track:"
                track_ws['B7'] = track_data['summary']['avg_streams_per_track']
                track_ws['B7'].font = Font(bold=True)
                track_ws['C7'] = service.format_number(track_data['summary']['avg_streams_per_track'])
                
                # Top track info
                if track_data.get('top_track'):
                    track_ws['A9'] = "Top Track:"
                    track_ws['B9'] = track_data['top_track']['track_name']
                    track_ws['B9'].font = Font(bold=True, color="FF0000")
                    track_ws['C9'] = service.format_number(track_data['top_track']['total_streams']) + " streams"
                
                # Track table headers
                track_row = 11
                track_headers = ['Track Name', 'Artist Credit', 'Total Streams', 'Avg Daily Streams', 'Peak Streams', 'Best Position', 'Weeks on Chart', 'Entry Date', 'Last Seen', 'Data Points']
                for col, header in enumerate(track_headers, 1):
                    cell = track_ws.cell(row=track_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Track data rows
                for idx, track in enumerate(track_data['tracks'], 1):
                    track_row += 1
                    track_ws.cell(row=track_row, column=1, value=track['track_name'])
                    track_ws.cell(row=track_row, column=2, value=track.get('track_credit', ''))
                    track_ws.cell(row=track_row, column=3, value=track['total_streams'])
                    track_ws.cell(row=track_row, column=4, value=round(track['avg_daily_streams'], 2))
                    track_ws.cell(row=track_row, column=5, value=track['peak_streams'])
                    track_ws.cell(row=track_row, column=6, value=track['best_position'])
                    track_ws.cell(row=track_row, column=7, value=track.get('weeks_on_chart', 'N/A'))
                    track_ws.cell(row=track_row, column=8, value=track.get('entry_date', 'N/A'))
                    track_ws.cell(row=track_row, column=9, value=track.get('last_appearance', 'N/A'))
                    track_ws.cell(row=track_row, column=10, value=track['data_points'])
                    
                    # Highlight top track (first row)
                    if idx == 1:
                        for col in range(1, 11):
                            track_ws.cell(row=track_row, column=col).fill = PatternFill(
                                start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"
                            )
                
                # Total row
                track_row += 1
                track_ws.cell(row=track_row, column=1, value="TOTAL")
                track_ws.cell(row=track_row, column=1).font = Font(bold=True)
                track_ws.cell(row=track_row, column=3, value=track_data['summary']['total_streams'])
                track_ws.cell(row=track_row, column=3).font = Font(bold=True)
                
                # Style total row
                for col in range(1, 9):
                    track_ws.cell(row=track_row, column=col).fill = PatternFill(
                        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                    )
                
                # Note
                track_row += 2
                track_ws.cell(row=track_row, column=1, value="Note:")
                track_ws.cell(row=track_row, column=2, value=track_data.get('note', ''))
                track_ws.cell(row=track_row, column=2).font = Font(italic=True, size=9)
                
                # Auto-adjust column widths for track sheet
                for column in track_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    track_ws.column_dimensions[column_letter].width = adjusted_width
                
                logger.info(f"Added track sheet for {artist['name']} on {platform['name']} with {len(track_data['tracks'])} tracks")
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"music_analytics_{start_date}_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except json.JSONDecodeError as e:
        logger.error(f"Excel export JSON decode error: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Invalid JSON data: {str(e)}'
        }, status=400)
    except ValueError as e:
        logger.error(f"Excel export value error: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Invalid data format: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Export failed: {str(e)}'
        }, status=500)